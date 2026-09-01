from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1E3A8A")


def _autosize_columns(ws, max_width: int = 42) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]]) -> int:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row_idx = start_row + 1
    for row in rows:
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1
    _autosize_columns(ws)
    return row_idx - 1


def _aggregate_by_course(students: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = defaultdict(lambda: {"students": 0, "balance": 0.0, "paid": 0.0, "total": 0.0})
    for s in students:
        course = str(s.get("course") or "Unknown")
        grouped[course]["students"] += 1
        grouped[course]["balance"] += float(s.get("balance") or 0)
        grouped[course]["paid"] += float(s.get("paid") or 0)
        grouped[course]["total"] += float(s.get("grand_total") or 0)
    return [
        {"course": course, **values}
        for course, values in sorted(grouped.items(), key=lambda x: x[1]["balance"], reverse=True)
    ]


def _aggregate_by_batch(students: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = defaultdict(lambda: {"students": 0, "balance": 0.0})
    for s in students:
        key = f"{s.get('course') or 'Unknown'} | {s.get('batch') or 'Unknown'}"
        grouped[key]["students"] += 1
        grouped[key]["balance"] += float(s.get("balance") or 0)
    return [
        {"batch_key": key, **values}
        for key, values in sorted(grouped.items(), key=lambda x: x[1]["balance"], reverse=True)
    ]


def build_pending_fee_workbook(students: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Students"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    headers = [
        "Reg No",
        "Student Name",
        "Course",
        "Batch / Session",
        "Total Fee",
        "Paid",
        "Balance",
        "Pending %",
    ]
    rows = [
        [
            s.get("student_reg_no"),
            s.get("student_name"),
            s.get("course"),
            s.get("batch"),
            float(s.get("grand_total") or 0),
            float(s.get("paid") or 0),
            float(s.get("balance") or 0),
            float(s.get("pending_percent") or 0),
        ]
        for s in students
    ]
    last_row = _write_table(ws, 4, headers, rows)

    summary = _aggregate_by_course(students)
    ws2 = wb.create_sheet("Course Summary")
    ws2["A1"] = "Course-wise Pending Fee Summary"
    ws2["A1"].font = TITLE_FONT
    summary_headers = ["Course", "Students", "Total Fee", "Paid", "Pending Balance"]
    summary_rows = [
        [
            item["course"],
            item["students"],
            round(item["total"], 2),
            round(item["paid"], 2),
            round(item["balance"], 2),
        ]
        for item in summary
    ]
    summary_last = _write_table(ws2, 3, summary_headers, summary_rows)

    if summary_rows:
        pie = PieChart()
        pie.title = "Pending Balance by Course"
        pie.height = 8
        pie.width = 14
        labels = Reference(ws2, min_col=1, min_row=4, max_row=summary_last)
        data = Reference(ws2, min_col=5, min_row=3, max_row=summary_last)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws2.add_chart(pie, "G3")

    batch_summary = _aggregate_by_batch(students)
    ws3 = wb.create_sheet("Batch Summary")
    ws3["A1"] = "Batch / Session-wise Pending Fee"
    ws3["A1"].font = TITLE_FONT
    batch_rows = [[b["batch_key"], b["students"], round(b["balance"], 2)] for b in batch_summary]
    batch_last = _write_table(ws3, 3, ["Course | Batch/Session", "Students", "Pending Balance"], batch_rows)
    if batch_rows:
        pie2 = PieChart()
        pie2.title = "Pending by Batch/Session"
        pie2.height = 8
        pie2.width = 14
        labels = Reference(ws3, min_col=1, min_row=4, max_row=batch_last)
        data = Reference(ws3, min_col=3, min_row=3, max_row=batch_last)
        pie2.add_data(data, titles_from_data=True)
        pie2.set_categories(labels)
        ws3.add_chart(pie2, "F3")

    ws.freeze_panes = "A5"
    return wb


def build_course_summary_workbook(courses: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Summary"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    headers = ["Course", "Students", "Assigned", "Collected", "Pending Balance"]
    rows = [
        [
            c.get("course"),
            int(c.get("students") or 0),
            float(c.get("assigned") or 0),
            float(c.get("collected") or 0),
            float(c.get("balance") or 0),
        ]
        for c in courses
    ]
    last_row = _write_table(ws, 4, headers, rows)

    if rows:
        pie = PieChart()
        pie.title = "Pending Balance by Course"
        pie.height = 9
        pie.width = 15
        labels = Reference(ws, min_col=1, min_row=5, max_row=last_row)
        data = Reference(ws, min_col=5, min_row=4, max_row=last_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "H4")

        bar = BarChart()
        bar.type = "col"
        bar.title = "Collected vs Pending by Course"
        bar.height = 9
        bar.width = 15
        collected = Reference(ws, min_col=4, min_row=4, max_row=last_row)
        pending = Reference(ws, min_col=5, min_row=4, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=5, max_row=last_row)
        bar.add_data(collected, titles_from_data=True)
        bar.add_data(pending, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "H20")

    return wb


def build_batch_summary_workbook(batches: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Summary"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    headers = ["Course", "Batch / Session", "Assigned", "Collected", "Pending Balance"]
    rows = [
        [
            b.get("course"),
            b.get("batch"),
            float(b.get("assigned") or 0),
            float(b.get("collected") or 0),
            float(b.get("balance") or 0),
        ]
        for b in batches
    ]
    last_row = _write_table(ws, 4, headers, rows)
    if rows:
        pie = PieChart()
        pie.title = "Pending by Batch/Session"
        pie.height = 9
        pie.width = 15
        labels = Reference(ws, min_col=2, min_row=5, max_row=last_row)
        data = Reference(ws, min_col=5, min_row=4, max_row=last_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "H4")
    return wb


def build_defaulter_workbook(defaulters: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Defaulters"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    headers = ["Student", "Reg No", "Course", "Batch", "Overdue Amount", "Max Days Overdue"]
    rows = [
        [
            d.get("student_name"),
            d.get("student_reg_no"),
            d.get("course"),
            d.get("batch"),
            float(d.get("overdue_amount") or 0),
            int(d.get("max_days_overdue") or 0),
        ]
        for d in defaulters
    ]
    last_row = _write_table(ws, 4, headers, rows)
    if rows:
        pie = PieChart()
        pie.title = "Overdue by Course"
        pie.height = 9
        pie.width = 15
        labels = Reference(ws, min_col=3, min_row=5, max_row=last_row)
        data = Reference(ws, min_col=5, min_row=4, max_row=last_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "H4")
    return wb


def build_daily_collection_workbook(collections: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Collection"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    headers = ["Date", "Payment Mode", "Receipts", "Amount"]
    rows = [
        [
            c.get("date"),
            c.get("mode"),
            int(c.get("receipts") or 0),
            float(c.get("amount") or 0),
        ]
        for c in collections
    ]
    last_row = _write_table(ws, 4, headers, rows)
    if rows:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Collection by Date"
        bar.height = 9
        bar.width = 15
        data = Reference(ws, min_col=4, min_row=4, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=5, max_row=last_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "G4")
    return wb


def save_workbook(wb: Workbook, directory: Path, filename: str) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / filename
    wb.save(path)
    return path
